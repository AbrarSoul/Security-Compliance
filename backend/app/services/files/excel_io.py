"""Read Excel workbooks (.xlsx) for metadata extraction and compliance scanning."""

import io
from typing import Any

from fastapi import HTTPException, status
from openpyxl import Workbook, load_workbook


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _rows_from_sheet(sheet) -> list[list[str]]:
    rows: list[list[str]] = []
    for row in sheet.iter_rows(values_only=True):
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
        rows.append([normalize_cell(cell) for cell in row])
    return rows


def read_xlsx_sheets(content: bytes) -> list[tuple[str, list[list[str]]]]:
    """Return non-empty sheets as (name, rows) pairs."""
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid or unreadable Excel workbook",
        ) from exc

    sheets: list[tuple[str, list[list[str]]]] = []
    try:
        for sheet_name in workbook.sheetnames:
            rows = _rows_from_sheet(workbook[sheet_name])
            if rows:
                sheets.append((sheet_name, rows))
    finally:
        workbook.close()

    return sheets


def validate_xlsx_content(content: bytes) -> None:
    if not content.startswith(b"PK"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid Excel workbook content",
        )
    read_xlsx_sheets(content)


def build_sample_xlsx(rows: list[list[str]], sheet_name: str = "Sheet1") -> bytes:
    """Build a minimal .xlsx workbook in memory (used by tests)."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for row in rows:
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
