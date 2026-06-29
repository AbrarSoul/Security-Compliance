#!/usr/bin/env python3
"""Generate 5 GAIRA use cases with answers for every framework module → Excel workbook."""

from __future__ import annotations

import json
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("Install openpyxl: pip install openpyxl")

ROOT = Path(__file__).resolve().parents[1]
FRAMEWORK_PATH = ROOT / "backend" / "app" / "data" / "gaira" / "framework_v1.json"
NIST_PATH = ROOT / "backend" / "app" / "data" / "nist_ai_rmf" / "controls_v1.json"
OUTPUT_PATH = ROOT / "GAIRA_USE_CASES_5_FRAMEWORKS.xlsx"

# ---------------------------------------------------------------------------
# Five realistic use cases
# ---------------------------------------------------------------------------
USE_CASES = [
    {
        "id": "UC1",
        "name": "Internal Meeting Summarizer",
        "owner": "Alex Chen",
        "department": "Operations",
        "purpose": "Summarize internal team meeting transcripts for action-item tracking",
        "audience": "Operations staff only (internal)",
        "ai_provider": "GPT-Lab (TUNI)",
        "technology_description": "Llama 3.1 8B local via GPT-Lab; batch upload of .txt transcripts; no external API; outputs stored in internal wiki",
        "compliance_model": "DEMO_LOCAL_LLM",
        "deployment": "local",
        "risk_tier": "low",
        "recommended_module": "gaira_light",
        "proceed_decision": "Proceed as planned",
        "area": "Internal productivity / operations",
        "interests": "Reduce manual note-taking; improve meeting follow-up consistency",
        "input_data": "Internal meeting transcripts (no customer PII); staff names may appear",
        "output_data": "Bullet summaries and action items in plain text",
        "frequency": "Daily after scheduled meetings",
        "training": "No fine-tuning; zero-shot prompting only",
        "data_subjects": "Employees (meeting participants)",
        "personal_data": "Minimal — staff names in transcripts",
        "external_parties": False,
        "automated_decisions": False,
        "high_risk_eu_act": False,
        "sensitive_processing": False,
        "customer_facing": False,
        "financial_impact": False,
        "health_advice": False,
        "credit_scoring": False,
    },
    {
        "id": "UC2",
        "name": "HR Policy Q&A Assistant",
        "owner": "Maria Kowalski",
        "department": "Human Resources",
        "purpose": "Answer employee questions about HR policies (leave, travel, benefits)",
        "audience": "All university employees",
        "ai_provider": "GPT-Lab (TUNI)",
        "technology_description": "Llama 3.1 8B local RAG over approved HR PDFs/CSVs; internal web chat; data stays on-premises",
        "compliance_model": "DEMO_LOCAL_LLM",
        "deployment": "local",
        "risk_tier": "medium",
        "recommended_module": "gaira_light",
        "proceed_decision": "Proceed with conditions",
        "area": "Human resources / employee services",
        "interests": "Faster HR self-service; consistent policy answers",
        "input_data": "Employee questions; HR policy documents in vector store",
        "output_data": "Policy guidance text with citations to source documents",
        "frequency": "Continuous during business hours",
        "training": "RAG index over approved HR docs; no model fine-tuning",
        "data_subjects": "Employees",
        "personal_data": "Employment-related queries may include personal circumstances",
        "external_parties": False,
        "automated_decisions": False,
        "high_risk_eu_act": False,
        "sensitive_processing": True,
        "customer_facing": False,
        "financial_impact": False,
        "health_advice": False,
        "credit_scoring": False,
    },
    {
        "id": "UC3",
        "name": "Customer Support Chatbot",
        "owner": "James Okonkwo",
        "department": "Customer Experience",
        "purpose": "Handle tier-1 product and billing questions on the public website",
        "audience": "External customers and prospects",
        "ai_provider": "OpenAI (Azure OpenAI)",
        "technology_description": "GPT-4 Turbo via Azure OpenAI West Europe; website widget; customer messages sent to external endpoint",
        "compliance_model": "DEMO_CLOUD_UNAPPROVED",
        "deployment": "external",
        "risk_tier": "medium",
        "recommended_module": "gaira_light",
        "proceed_decision": "Proceed with conditions",
        "area": "Customer service / public-facing digital channels",
        "interests": "Reduce support queue; 24/7 first-line responses",
        "input_data": "Customer questions, account email, order references",
        "output_data": "Support answers, escalation links, ticket summaries",
        "frequency": "Continuous (24/7)",
        "training": "RAG over public FAQ and product docs; Azure content filtering enabled",
        "data_subjects": "Customers, prospects",
        "personal_data": "Names, emails, order IDs, billing questions",
        "external_parties": True,
        "automated_decisions": False,
        "high_risk_eu_act": False,
        "sensitive_processing": True,
        "customer_facing": True,
        "financial_impact": True,
        "health_advice": False,
        "credit_scoring": False,
    },
    {
        "id": "UC4",
        "name": "Loan Pre-Approval Scoring",
        "owner": "Elena Vasquez",
        "department": "Retail Banking",
        "purpose": "Pre-screen loan applications and rank creditworthiness for human underwriter review",
        "audience": "Loan applicants (consumers)",
        "ai_provider": "Internal ML Platform",
        "technology_description": "Proprietary gradient-boosting model + LLM explanation layer; hosted in bank VPC; integrates with core banking API",
        "compliance_model": "DEMO_EXTERNAL_API",
        "deployment": "external",
        "risk_tier": "high",
        "recommended_module": "gaira_comprehensive",
        "proceed_decision": "Proceed with conditions",
        "area": "Financial services / credit decision support",
        "interests": "Faster loan processing; consistent risk scoring",
        "input_data": "Applicant income, credit history, employment, application form data",
        "output_data": "Risk score, recommendation tier, explanation summary for underwriter",
        "frequency": "Per loan application (batch + real-time)",
        "training": "Fine-tuned on historical anonymized loan performance data",
        "data_subjects": "Loan applicants (consumers)",
        "personal_data": "Financial and identity data; special categories possible",
        "external_parties": True,
        "automated_decisions": True,
        "high_risk_eu_act": True,
        "sensitive_processing": True,
        "customer_facing": True,
        "financial_impact": True,
        "health_advice": False,
        "credit_scoring": True,
    },
    {
        "id": "UC5",
        "name": "Clinical Symptom Triage Chatbot",
        "owner": "Dr. Sami Rahman",
        "department": "Digital Health",
        "purpose": "Triage patient symptoms and suggest urgency level before human clinician contact",
        "audience": "Patients using regional health portal",
        "ai_provider": "MedAI Cloud",
        "technology_description": "MedAI clinical LLM via BAA-covered cloud; patient portal integration; PHI processed under healthcare contract",
        "compliance_model": "DEMO_EXTERNAL_API",
        "deployment": "external",
        "risk_tier": "very_high",
        "recommended_module": "gaira_comprehensive",
        "proceed_decision": "Further review required",
        "area": "Healthcare / patient-facing triage",
        "interests": "Reduce ER overload; faster routing to appropriate care",
        "input_data": "Patient-reported symptoms, age band, medications, medical history snippets",
        "output_data": "Urgency recommendation, care pathway, disclaimer to seek emergency care",
        "frequency": "Continuous patient self-service",
        "training": "Vendor pre-trained clinical model; local fine-tune on regional protocols",
        "data_subjects": "Patients (health data subjects)",
        "personal_data": "Health data (special category); identifiers",
        "external_parties": True,
        "automated_decisions": True,
        "high_risk_eu_act": True,
        "sensitive_processing": True,
        "customer_facing": True,
        "financial_impact": True,
        "health_advice": True,
        "credit_scoring": False,
    },
]

UC_IDS = [uc["id"] for uc in USE_CASES]
UC_NAMES = {uc["id"]: uc["name"] for uc in USE_CASES}

# High-risk question IDs in ai_risk_levels (step 2)
HIGH_RISK_QS = {f"2.{i:02d}" for i in range(1, 11)}
# Medium-risk question IDs (step 3)
MEDIUM_RISK_QS = {f"3.{i:02d}" for i in range(1, 15)}
# Insignificant (step 1)
INSIG_QS = {"1.01", "1.03", "1.04"}


def ai_risk_levels_answer(qid: str, uc: dict) -> str:
    """Yes/No answers following GAIRA triage logic."""
    tier = uc["risk_tier"]

    if tier in ("low", "insignificant"):
        if qid in INSIG_QS:
            return "Yes" if qid == "1.01" else "No"
        if qid in HIGH_RISK_QS:
            return "No"
        if qid in MEDIUM_RISK_QS:
            return "Yes" if qid in ("3.01", "3.06") and uc.get("sensitive_processing") else "No"
        return "No"

    if tier == "medium":
        if qid in INSIG_QS:
            return "No"
        if qid in HIGH_RISK_QS:
            return "No"
        if qid in MEDIUM_RISK_QS:
            if qid == "3.03" and uc.get("external_parties"):
                return "Yes"
            if qid == "3.06" and uc.get("sensitive_processing"):
                return "Yes"
            if qid == "3.05" and uc.get("financial_impact"):
                return "Yes"
            if qid == "3.04" and uc.get("customer_facing"):
                return "Yes"
            if qid == "3.01" and uc.get("automated_decisions"):
                return "Yes"
            return "No"
        return "No"

    # high / very_high
    if qid in INSIG_QS:
        return "No"
    if qid in HIGH_RISK_QS:
        if qid == "2.02" and uc.get("automated_decisions"):
            return "Yes"
        if qid == "2.03" and (uc.get("health_advice") or uc.get("credit_scoring")):
            return "Yes"
        if qid == "2.06" and uc.get("high_risk_eu_act"):
            return "Yes"
        if qid == "2.07" and uc.get("financial_impact"):
            return "Yes"
        if qid == "2.08" and uc.get("credit_scoring"):
            return "Yes"
        if qid == "2.05" and uc.get("customer_facing"):
            return "Yes"
        if qid == "2.09":
            return "Yes"
        if qid == "2.10":
            return "Yes"
        return "No"
    if qid in MEDIUM_RISK_QS:
        return "Yes" if qid in ("3.01", "3.02", "3.05", "3.06", "3.04") else "No"
    return "No"


def gaira_section1_answer(qid: str, uc: dict) -> str:
    mapping = {
        "1.01": uc["area"],
        "1.02": uc["purpose"],
        "1.03": uc["interests"],
        "1.04": uc["technology_description"],
        "1.05": f"{uc['ai_provider']} — model provider; internal IT — integration and hosting",
        "1.06": uc["input_data"],
        "1.07": uc["output_data"],
        "1.08": uc["frequency"],
        "1.09": uc["training"],
        "1.10": uc["data_subjects"],
        "1.11": uc["personal_data"],
        "1.12": "Internal staff only" if not uc["external_parties"] else "External customers/patients/applicants",
        "1.13": "No special categories" if not uc["health_advice"] and not uc["credit_scoring"] else (
            "Health data (Art. 9 GDPR)" if uc["health_advice"] else "Financial data; possible special categories"
        ),
        "1.14": "Local/on-prem" if uc["deployment"] == "local" else "Cloud (contractual safeguards required)",
        "1.15": uc["owner"],
        "1.16": "Documented in ROAIA; reviewed quarterly" if uc["risk_tier"] in ("high", "very_high") else "Documented in ROAIA",
        "1.17": "Aligned with group AI policy v2.1",
        "1.18": "Deployer" if uc["customer_facing"] else "Internal deployer",
        "1.19": "High-risk" if uc["high_risk_eu_act"] else "Not high-risk / limited risk",
        "1.20": "Efficiency, consistency, improved user experience",
    }
    return mapping.get(qid, uc["purpose"])


def yes_no_text(condition: bool, yes: str = "Yes", no: str = "No") -> str:
    return yes if condition else no


def gaira_step2_answer(qid: str, uc: dict) -> str:
    """EU AI Act / scope questions (text, typically Yes/No + brief note)."""
    answers = {
        "2.01": yes_no_text(True, "Yes — LLM/ML generates output from prompts without fully predefined rules"),
        "2.02": yes_no_text(uc["health_advice"], "No", "No — not prohibited practices under Art. 5"),
        "2.03": yes_no_text(uc["high_risk_eu_act"], "Yes — Annex III high-risk (credit/health)" if uc["high_risk_eu_act"] else "No"),
        "2.04": yes_no_text(not uc["high_risk_eu_act"], "Yes — narrow task / human oversight", "No — high-risk path applies"),
        "2.05": yes_no_text(uc["risk_tier"] in ("low", "medium") and not uc["credit_scoring"], "No", "No"),
        "2.06": yes_no_text(uc["deployment"] == "local", "Partially — integrated locally", "No — third-party vendor model"),
        "2.07": yes_no_text(uc["external_parties"], "Yes — external API provider", "No — internal only"),
        "2.08": yes_no_text(True, "Yes — documented roles in vendor DPA"),
        "2.09": yes_no_text(uc["customer_facing"], "Yes", "No"),
        "2.10": yes_no_text(uc["high_risk_eu_act"], "Yes — conformity assessment planned", "No"),
        "2.11": yes_no_text(uc["automated_decisions"], "Yes — human in the loop for final decision", "No — advisory only"),
        "2.12": yes_no_text(uc["sensitive_processing"], "Yes — DPIA / transfer impact assessment", "No"),
        "2.13": yes_no_text(uc["deployment"] == "external", "Yes — SCCs / BAA in place", "No — data stays local"),
        "2.14": yes_no_text(True, "Yes — logging and retention per policy"),
        "2.15": yes_no_text(uc["customer_facing"], "Yes — visible disclaimer", "N/A — internal tool"),
        "2.16": yes_no_text(uc["risk_tier"] in ("high", "very_high"), "Yes — incident response playbook", "Standard IT incident process"),
        "2.17": yes_no_text(uc["high_risk_eu_act"], "Yes — bias testing for credit/health", "Basic QA sampling"),
        "2.18": yes_no_text(True, "Yes — version controlled prompts/models"),
    }
    return answers.get(qid, "See application profile in ROAIA")


def gaira_risk_step_answer(qid: str, step: str, uc: dict, module: str) -> str:
    """Steps 3–6: risk controls, mitigations, flags."""
    tier = uc["risk_tier"]
    if step == "3":
        mitigations = {
            "3.01": "Human review before any external action" if uc["automated_decisions"] else "Output reviewed by user before sharing",
            "3.02": "Access limited to authorized roles; MFA enforced",
            "3.03": "Encryption at rest and in transit; secrets vault",
            "3.04": "Prompt injection filters; guardrails in ComplianceGuard",
            "3.05": "Approved model registry only; block external models for sensitive data",
            "3.06": "Data minimization; PII masking in datasets",
            "3.07": "Audit logging of all prompts and outputs",
            "3.08": "Vendor SLA and fallback to human process",
            "3.09": "Quarterly access review",
            "3.10": "Staff training on acceptable use",
        }
        num = qid.split(".")[1] if "." in qid else ""
        return mitigations.get(qid, f"Mitigation documented for {uc['name']}; owner: {uc['owner']}")

    if step == "4":
        # Problematic / compliance checkpoints — flag high-risk items
        flags = {
            "4.01": yes_no_text(uc["deployment"] == "external", "Yes — data leaves platform", "No"),
            "4.02": yes_no_text(uc["sensitive_processing"], "Yes", "No"),
            "4.03": yes_no_text(not uc.get("high_risk_eu_act", False) or tier == "medium", "Adequate", "Requires legal review"),
            "4.04": yes_no_text(tier in ("low", "medium"), "Adequate", "Enhanced controls required"),
            "4.05": yes_no_text(uc["customer_facing"], "Yes — public channel", "No"),
            "4.10": yes_no_text(uc["health_advice"] or uc["credit_scoring"], "Yes — 2nd line review", "No"),
            "4.20": yes_no_text(uc["deployment"] == "local", "Yes — logging enabled", "No — verify vendor logs"),
        }
        if qid in flags:
            return flags[qid]
        return yes_no_text(tier in ("low", "medium"), "No issues identified", "Review required — see risk register")

    if step == "5":
        if qid.endswith(".01") or "risk" in qid.lower():
            return uc["risk_tier"].replace("_", " ").title()
        if "proceed" in qid.lower() or qid in ("5.01", "6.01"):
            return uc["proceed_decision"]
        return f"Owner assessment for {uc['name']}: acceptable with documented controls"

    if step == "6":
        if qid in ("6.01", "6.02"):
            return uc["proceed_decision"]
        if "risk" in qid.lower():
            risk_map = {"low": "Low", "medium": "Medium", "high": "High", "very_high": "Very high", "insignificant": "Low"}
            return risk_map.get(tier, "Medium")
        return uc["owner"]

    return f"Completed per {module} guidance for {uc['id']}"


def answer_text_question(q: dict, uc: dict, module: str) -> str:
    qid = q["id"]
    step = q["step_id"]

    if qid in {f"1.{i:02d}" for i in range(1, 21)}:
        return gaira_section1_answer(qid, uc)

    if step == "2":
        return gaira_step2_answer(qid, uc)

    if module in ("gaira_light", "gaira_comprehensive") and step in ("3", "4", "5", "6"):
        return gaira_risk_step_answer(qid, step, uc, module)

    if module == "ai_act_check":
        return ai_act_answer(qid, q["text"], uc)

    if module == "compliance_check":
        return compliance_check_answer(qid, q["text"], uc)

    return uc["purpose"]


def ai_act_answer(qid: str, text: str, uc: dict) -> str:
    t = text.lower()
    if "provider" in t or "deployer" in t or "role" in t:
        return "Deployer" if not uc["credit_scoring"] else "Deployer (financial institution)"
    if "prohibited" in t or "art. 5" in t:
        return "No — not a prohibited practice"
    if "high-risk" in t or "annex" in t:
        return "Yes — Annex III applies" if uc["high_risk_eu_act"] else "No — not classified as high-risk"
    if "gpai" in t or "general-purpose" in t:
        return "No — application-specific deployment"
    if "transparency" in t or "disclosure" in t:
        return "Yes — users informed they interact with AI" if uc["customer_facing"] else "N/A — internal"
    if "conformity" in t or "ce marking" in t:
        return "Planned before production" if uc["high_risk_eu_act"] else "Not required"
    if "fundamental rights" in t:
        return "FRIA planned" if uc["high_risk_eu_act"] else "Standard DPIA where personal data processed"
    if "scope" in t or "apply" in t:
        return "Yes — EU operations / EU data subjects affected"
    return f"Assessed for {uc['name']}: documented in AI Act worksheet"


def compliance_check_answer(qid: str, text: str, uc: dict) -> str:
    t = text.lower()
    if "dpia" in t or "impact assessment" in t:
        return "Required — in progress" if uc["sensitive_processing"] else "Not required"
    if "legal basis" in t or "lawful" in t:
        return "Legitimate interest / contract" if uc["customer_facing"] else "Legitimate interest (internal)"
    if "retention" in t:
        return "90 days logs; 2 years audit" if uc["risk_tier"] in ("high", "very_high") else "30 days prompts; 1 year audit"
    if "encryption" in t:
        return "AES-256 at rest; TLS 1.3 in transit"
    if "access" in t or "rbac" in t:
        return "Role-based access; least privilege; MFA for admins"
    if "vendor" in t or "third" in t:
        return "DPA + SCCs signed" if uc["deployment"] == "external" else "N/A — internal hosting"
    if "incident" in t:
        return "72h breach notification procedure documented"
    if "training" in t and "staff" in t:
        return "Annual AI acceptable-use training completed"
    if "monitor" in t or "audit" in t:
        return "ComplianceGuard scans + execution guard + audit logs enabled"
    if "bias" in t or "fair" in t:
        return "Bias testing scheduled" if uc["credit_scoring"] else "Periodic output sampling"
    return f"Compliant / documented for {uc['name']}"


def computed_risk_level(uc: dict) -> str:
    return {
        "insignificant": "Insignificant",
        "low": "Low",
        "medium": "Medium",
        "high": "High",
        "very_high": "Very high",
    }.get(uc["risk_tier"], "Medium")


def nist_status_for_uc(control: dict, uc: dict) -> str:
    """Scenario-level NIST control posture if this use case were the primary deployment."""
    ev = control.get("evaluator")
    tier = uc["risk_tier"]
    modules = control.get("modules") or []

    if control.get("coverage") == "none":
        return "not_applicable"

    if "gaira" in modules and tier in ("high", "very_high"):
        return "partial" if ev else "met"
    if tier in ("low",) and ev in ("org_has_active_controls",):
        return "met"
    if tier == "medium":
        return "partial" if ev else "not_assessed"
    if tier in ("high", "very_high"):
        if ev in ("gaira_assessments_complete", "execution_guard_active"):
            return "partial"
        return "not_met" if ev else "partial"
    return control.get("coverage", "partial") or "partial"


def posture_framework_status(framework: str, uc: dict) -> tuple[str, str]:
    tier = uc["risk_tier"]
    if framework == "gaira":
        if tier in ("low", "insignificant"):
            return "met", "GAIRA Light completed; ROAIA updated"
        if tier == "medium":
            return "partial", "GAIRA Light done; conditions open (monitoring, data minimization)"
        return "partial", "Comprehensive assessment required; 2nd line review pending"
    if framework == "nist_ai_rmf":
        if tier in ("low", "medium"):
            return "partial", "Core controls met; formal AI RMF profile partially evidenced"
        return "not_met", "High-risk use case — multiple NIST controls need strengthening"
    if framework == "internal_guardrails":
        if uc["deployment"] == "local" and tier in ("low", "medium"):
            return "met", "Local model, active policies, scans before execution"
        if uc["deployment"] == "external":
            return "partial", "External model — warn/block rules must be active; DPA required"
        return "not_met", "Critical gaps likely until comprehensive controls deployed"
    return "not_assessed", ""


def style_header(ws, row: int = 1):
    fill = PatternFill("solid", fgColor="1E3A5F")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def auto_width(ws, max_width: int = 48):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[letter].width = min(max(length + 2, 12), max_width)


def build_overview_sheet(wb: Workbook):
    ws = wb.active
    ws.title = "Use Cases Overview"
    headers = [
        "Use Case ID",
        "Application Name",
        "Owner",
        "Department",
        "Purpose",
        "Audience",
        "AI Provider",
        "Technology / Model",
        "Compliance Model",
        "Deployment",
        "Expected Risk Tier",
        "Recommended GAIRA Path",
        "Proceed Decision",
        "AI Risk Levels Result",
        "GAIRA Module Used",
        "Also Complete",
    ]
    ws.append(headers)
    for uc in USE_CASES:
        also = ["ai_act_check", "compliance_check"]
        if uc["risk_tier"] in ("high", "very_high"):
            module = "gaira_comprehensive"
        else:
            module = "gaira_light"
        ws.append(
            [
                uc["id"],
                uc["name"],
                uc["owner"],
                uc["department"],
                uc["purpose"],
                uc["audience"],
                uc["ai_provider"],
                uc["technology_description"],
                uc["compliance_model"],
                uc["deployment"],
                computed_risk_level(uc),
                uc["recommended_module"],
                uc["proceed_decision"],
                computed_risk_level(uc),
                module,
                ", ".join(also),
            ]
        )
    style_header(ws)
    auto_width(ws, 60)


def build_module_sheet(wb: Workbook, module_key: str, framework: dict):
    title = module_key.replace("_", " ").title()[:31]
    ws = wb.create_sheet(title)
    mod = framework["modules"][module_key]
    headers = ["Module", "Step", "Question ID", "Question Text"] + [f"{uid} — {UC_NAMES[uid]}" for uid in UC_IDS]
    ws.append(headers)

    for q in mod["questions"]:
        row = [
            module_key,
            q.get("step_id", ""),
            q["id"],
            q["text"],
        ]
        for uc in USE_CASES:
            if module_key == "ai_risk_levels":
                row.append(ai_risk_levels_answer(q["id"], uc))
            else:
                row.append(answer_text_question(q, uc, module_key))
        ws.append(row)

    # Select questions at end of gaira_light
    if module_key == "gaira_light":
        ws.append([])
        ws.append(["", "", "", "Overall risk level (select)"] + [computed_risk_level(uc) for uc in USE_CASES])
        ws.append(["", "", "", "Proceed decision (select)"] + [uc["proceed_decision"] for uc in USE_CASES])

    style_header(ws)
    auto_width(ws, 55)


def build_nist_sheet(wb: Workbook, nist_data: dict):
    ws = wb.create_sheet("NIST AI RMF")
    headers = ["Control ID", "Function", "Control Text", "Evaluator"] + [
        f"{uid} — scenario posture" for uid in UC_IDS
    ]
    ws.append(headers)
    for ctrl in nist_data["controls"]:
        row = [ctrl["id"], ctrl["function"], ctrl["text"], ctrl.get("evaluator") or ""]
        for uc in USE_CASES:
            row.append(nist_status_for_uc(ctrl, uc))
        ws.append(row)
    style_header(ws)
    auto_width(ws, 50)


def build_posture_sheet(wb: Workbook):
    ws = wb.create_sheet("Compliance Posture")
    frameworks = [
        ("nist_ai_rmf", "NIST AI RMF"),
        ("gaira", "GAIRA Governance"),
        ("internal_guardrails", "Internal Guardrails"),
    ]
    headers = ["Framework ID", "Framework Name"] + [f"{uid} — status" for uid in UC_IDS] + [
        f"{uid} — notes" for uid in UC_IDS
    ]
    ws.append(headers)
    for fid, fname in frameworks:
        row = [fid, fname]
        notes = []
        for uc in USE_CASES:
            status, _ = posture_framework_status(fid, uc)
            row.append(status)
        for uc in USE_CASES:
            _, note = posture_framework_status(fid, uc)
            notes.append(note)
        row.extend(notes)
        ws.append(row)
    style_header(ws)
    auto_width(ws, 40)


def main():
    framework = json.loads(FRAMEWORK_PATH.read_text(encoding="utf-8"))
    nist_data = json.loads(NIST_PATH.read_text(encoding="utf-8"))

    wb = Workbook()
    build_overview_sheet(wb)

    for module_key in [
        "ai_risk_levels",
        "gaira_light",
        "gaira_comprehensive",
        "ai_act_check",
        "compliance_check",
    ]:
        build_module_sheet(wb, module_key, framework)

    build_nist_sheet(wb, nist_data)
    build_posture_sheet(wb)

    wb.save(OUTPUT_PATH)
    print(f"Created: {OUTPUT_PATH}")
    print(f"Use cases: {len(USE_CASES)}")
    for key in framework["modules"]:
        n = len(framework["modules"][key].get("questions", []))
        if n:
            print(f"  {key}: {n} questions x {len(USE_CASES)} use cases")
    print(f"  NIST AI RMF: {len(nist_data['controls'])} controls x {len(USE_CASES)} scenarios")


if __name__ == "__main__":
    main()
